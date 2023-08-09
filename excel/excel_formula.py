"""
Create formula in spreadsheet/excel using python
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# to load and select worksheet
wb = load_workbook('barchat.xlsx')
worksheet = wb['Report']

# select the min_col, max_col, min_row, max_row, 
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# for one column
"""
worksheet['B8'] = 'SUM(B6:B7)'
worksheet['B8'].style = 'Currency'
"""

# for multiple columns
for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    worksheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    worksheet[f'{letter}{max_row+1}'].style = 'Currency'



wb.save('report.xlsx')
