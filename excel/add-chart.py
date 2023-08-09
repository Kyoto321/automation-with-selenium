"""
To create charts in excel using python

pip install openpyxl
"""
from openpyxl import load_workbook
from openpyxl import BarChart, Reference

# to load and select worksheet
wb = load_workbook('pivot_table.xlsx')
worksheet = wb['Report']

# select the min_col, max_col, min_row, max_row, 
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# to plot graph
barchart = BarChart()
# indicate the worksheet, min and max value(of row&col)
# # split worksheet into data and categories
data = Reference(worksheet,
          min_col = min_column+1,
          max_col = max_column,
          min_row = min_row,
          max_row = max_row)

categories = Reference(worksheet,
          min_col = min_column,
          max_col = min_column,
          min_row = min_row+1,
          max_row = max_row)

barchart.add_data(data, title_from_data=True)
barchart.add_data(categories)

worksheet.add_chart(barchart, "B12")


barchart.title = "Sales by product line"
barchart.style = 5
wb.save('barchart.xlsx')




