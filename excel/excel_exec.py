"""
Convert excel worksheeo into executable py file
"""
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl import BarChart, Reference
import os
import sys

#specified path
application_path = os.path.dirname(sys.executable)

# to set the current name of the month
month = input('Introduce month: ')

# to load and select worksheet
input_path = os.path.join(application_path, 'pivot_table.xlsx') # to put both files in thesame folder
wb = load_workbook(input_path)
worksheet = wb['Report']

# select the min_col, max_col, min_row, max_row, 
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# to plot graph
barchart = BarChart()
# indicate the worksheet, min and max value(of row&col)
# # split worksheet into data and categories
data = Reference(worksheet,
          min_col = min_column+1,
          max_col = max_column,
          min_row = min_row,
          max_row = max_row)

categories = Reference(worksheet,
          min_col = min_column,
          max_col = min_column,
          min_row = min_row+1,
          max_row = max_row)

# add data and categories
barchart.add_data(data, title_from_data=True)
barchart.add_data(categories)

# make chart
worksheet.add_chart(barchart, "B12")
barchart.title = "Sales by product line"
barchart.style = 5

# loop throuh columns with data to get total sum of each column
# # for multiple columns
for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    worksheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    worksheet[f'{letter}{max_row+1}'].style = 'Currency'

worksheet[f'{get_column_letter(min_column)}{max_row+1}'] = 'Total'

# worksheet report
worksheet['A1'] = 'Sales Report'
worksheet['A2'] = month
worksheet['A1'].font = Font('Arial', bold=True, size=20)
worksheet['A2'].font = Font('Arial', bold=True, size=10)

output_path = os.path.join(application_path, f'report_{month}.xlsx')

wb.save(output_path)















